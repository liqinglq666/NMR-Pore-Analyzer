"""
Professional multi-sheet Excel research report for NMR-Pore-Analyzer.

Workbook structure
------------------
1. Summary_Peak_Statistics
2. Pore_Classification_Ratios
3. Cumulative_Curve_Data
4. Differential_Curve_Data
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from logic.analyzer import AnalysisResult
from logic.peak_processor import PeakResult


# ---------------------------------------------------------------------------
# Shared style constants
# ---------------------------------------------------------------------------
_TNR = "Times New Roman"
_FONT = Font(name=_TNR, size=11)
_FONT_B = Font(name=_TNR, size=11, bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THICK = Side(border_style="medium", color="000000")
_THIN = Side(border_style="thin", color="000000")
_NONE = Side(border_style=None)


def _border(
    top: Side = _NONE,
    bottom: Side = _NONE,
    left: Side = _NONE,
    right: Side = _NONE,
) -> Border:
    return Border(top=top, bottom=bottom, left=left, right=right)


_B_TOP_THICK = _border(top=_THICK)
_B_BOT_THIN = _border(bottom=_THIN)
_B_BOT_THICK = _border(bottom=_THICK)
_B_NONE = _border()


def _fmt_r(value: float) -> str:
    """Format pore radius: one decimal if <1000 nm, else integer."""
    return f"{value:.1f}" if value < 1000 else str(round(value))


def _set_cell(
    ws,
    row: int,
    col: int,
    value,
    font: Font = _FONT,
    align: Alignment = _CENTER,
    border: Border = _B_NONE,
) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.alignment = align
    c.border = border


def _safe_sheet_title(title: str) -> str:
    """Excel sheet names are limited to 31 characters."""
    return title[:31]


# ---------------------------------------------------------------------------
# Sheet 1 - Summary_Peak_Statistics
# ---------------------------------------------------------------------------
_PS_GROUPS = [
    (1, 1, "Mixture"),
    (2, 4, "Main Peak"),
    (5, 7, "Sub-Peak"),
    (8, 8, "Boundary"),
    (9, 9, "Peak Ratio\n(Main / Sub)"),
]
_PS_SUB = [
    "Mixture",
    "T2 (ms)", "r (nm)", "Area (%)",
    "T2 (ms)", "r (nm)", "Area (%)",
    "Valley T2 (ms)",
    "Main / Sub",
]


def _write_peak_statistics_sheet(
    ws,
    batch: Dict[str, Tuple[AnalysisResult, PeakResult]],
) -> None:
    for col_s, col_e, label in _PS_GROUPS:
        if col_e > col_s:
            ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)
        _set_cell(ws, 1, col_s, label, font=_FONT_B, border=_B_TOP_THICK)
        for col in range(col_s + 1, col_e + 1):
            ws.cell(row=1, column=col).border = _B_TOP_THICK

    for col, label in enumerate(_PS_SUB, start=1):
        _set_cell(ws, 2, col, label, font=_FONT_B, border=_B_BOT_THIN)

    names = list(batch.keys())
    n = len(names)
    for i, name in enumerate(names):
        row = i + 3
        _an, pk = batch[name]
        bdr = _B_BOT_THICK if i == n - 1 else _B_NONE

        def _w(col: int, value, _row: int = row, _bdr: Border = bdr) -> None:
            _set_cell(ws, _row, col, value, border=_bdr)

        _w(1, name)
        _w(2, round(pk.primary.t2_ms, 2))
        _w(3, _fmt_r(pk.primary.radius_nm))
        _w(4, round(pk.primary.area_ratio * 100, 2))

        if pk.has_secondary and pk.secondary is not None and pk.valley is not None:
            _w(5, round(pk.secondary.t2_ms, 2))
            _w(6, _fmt_r(pk.secondary.radius_nm))
            _w(7, round(pk.secondary.area_ratio * 100, 2))
            _w(8, round(pk.valley.t2_ms, 2))
            ratio = (
                round(pk.primary.area_ratio / pk.secondary.area_ratio, 2)
                if pk.secondary.area_ratio > 0 else "N/A"
            )
            _w(9, ratio)
        else:
            for col in range(5, 10):
                _w(col, "N/A")

    for col, w in enumerate([14, 11, 11, 11, 11, 11, 11, 18, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


# ---------------------------------------------------------------------------
# Sheet 2 - Pore_Classification_Ratios
# ---------------------------------------------------------------------------
_A_LABELS = ["Gel", "Transition", "Capillary", "Air-voids"]
_B_LABELS = ["Harmless", "Less-harmful", "Harmful", "More-harmful"]
_PCR_GROUPS = [
    (1, 1, "Mixture"),
    (2, 5, "System A  (Gel / Capillary)"),
    (6, 9, "System B  (Harmless / Harmful)"),
]


def _write_pore_classification_sheet(
    ws,
    batch: Dict[str, Tuple[AnalysisResult, PeakResult]],
) -> None:
    for col_s, col_e, label in _PCR_GROUPS:
        if col_e > col_s:
            ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)
        _set_cell(ws, 1, col_s, label, font=_FONT_B, border=_B_TOP_THICK)
        for col in range(col_s + 1, col_e + 1):
            ws.cell(row=1, column=col).border = _B_TOP_THICK

    for col, label in enumerate(["Mixture"] + _A_LABELS + _B_LABELS, start=1):
        _set_cell(ws, 2, col, label, font=_FONT_B, border=_B_BOT_THIN)

    names = list(batch.keys())
    n = len(names)
    for i, name in enumerate(names):
        row = i + 3
        an, _pk = batch[name]
        bdr = _B_BOT_THICK if i == n - 1 else _B_NONE
        _set_cell(ws, row, 1, name, border=bdr)

        a_dict = dict(zip(an.system_a.labels, an.system_a.ratios))
        b_dict = dict(zip(an.system_b.labels, an.system_b.ratios))
        for j, lbl in enumerate(_A_LABELS, start=2):
            _set_cell(ws, row, j, round(a_dict.get(lbl, 0.0) * 100, 2), border=bdr)
        for j, lbl in enumerate(_B_LABELS, start=6):
            _set_cell(ws, row, j, round(b_dict.get(lbl, 0.0) * 100, 2), border=bdr)

    ws.column_dimensions["A"].width = 14
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22


# ---------------------------------------------------------------------------
# Curve sheets
# ---------------------------------------------------------------------------

def _differential_y(an: AnalysisResult) -> np.ndarray:
    """Amplitude as percentage of the total positive signal."""
    total = float(an.raw.amplitude.sum())
    if total <= 0:
        return np.zeros_like(an.raw.amplitude, dtype=float)
    return (an.raw.amplitude / total) * 100.0


def _write_curve_pair_sheet(
    ws,
    batch: Dict[str, Tuple[AnalysisResult, PeakResult]],
    *,
    y_header: str,
    y_unit: str,
    y_getter,
) -> None:
    """Write Origin-ready paired columns: Radius + Y per mixture."""
    for k, (name, (an, _pk)) in enumerate(batch.items()):
        c0 = 2 * k + 1
        c1 = c0 + 1

        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        _set_cell(ws, 1, c0, name, font=_FONT_B, border=_B_TOP_THICK)
        ws.cell(row=1, column=c1).border = _B_TOP_THICK

        _set_cell(ws, 2, c0, "Radius (nm)", font=_FONT_B, border=_B_BOT_THIN)
        _set_cell(ws, 2, c1, f"{y_header} ({y_unit})", font=_FONT_B, border=_B_BOT_THIN)

        x_vals = an.radius_nm
        y_vals = y_getter(an)
        n_rows = len(x_vals)
        for offset, (xv, yv) in enumerate(zip(x_vals, y_vals)):
            row = offset + 3
            bdr = _B_BOT_THICK if offset == n_rows - 1 else _B_NONE
            _set_cell(ws, row, c0, float(round(xv, 4)), border=bdr)
            _set_cell(ws, row, c1, float(round(float(yv), 6)), border=bdr)

    for k in range(len(batch)):
        c0 = 2 * k + 1
        ws.column_dimensions[get_column_letter(c0)].width = 14
        ws.column_dimensions[get_column_letter(c0 + 1)].width = 18
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_batch_results(
    output_path: Path,
    batch: Dict[str, Tuple[AnalysisResult, PeakResult]],
) -> Path:
    """Write a 4-sheet research workbook for all mixtures in batch."""
    if not batch:
        raise ValueError("No analysis results to export.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws1 = wb.create_sheet(_safe_sheet_title("Summary_Peak_Statistics"))
    _write_peak_statistics_sheet(ws1, batch)

    ws2 = wb.create_sheet(_safe_sheet_title("Pore_Classification_Ratios"))
    _write_pore_classification_sheet(ws2, batch)

    ws3 = wb.create_sheet(_safe_sheet_title("Cumulative_Curve_Data"))
    _write_curve_pair_sheet(
        ws3,
        batch,
        y_header="Cumulative porosity",
        y_unit="%",
        y_getter=lambda an: an.cumulative * 100.0,
    )

    ws4 = wb.create_sheet(_safe_sheet_title("Differential_Curve_Data"))
    _write_curve_pair_sheet(
        ws4,
        batch,
        y_header="Incremental signal fraction",
        y_unit="%",
        y_getter=_differential_y,
    )

    wb.save(output_path)
    return output_path.resolve()


def export_results(
    output_path: Path,
    analysis: AnalysisResult,
    peak_result: PeakResult,
) -> Path:
    """Convenience wrapper: export a single sample as a one-mixture batch."""
    return export_batch_results(
        output_path,
        batch={analysis.raw.sample_name: (analysis, peak_result)},
    )
